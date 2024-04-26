#importando a biblioteca de leitura de planilha 
from openpyxl import load_workbook

#selecionando a planilha e nome de pagina 
planilha = load_workbook('./fornecedores.xlsx')
pagina_planilha = planilha['Sheet1']

#lendo a planilha linha a linha 
for linha in pagina_planilha.iter_rows(min_row=2, values_only=True):
    nome_empresa, endereco, cidade, estado, cep, telefone, email, setor  = linha
    print(nome_empresa)
